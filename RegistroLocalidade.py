# pyinstaller RegistroLocalidade.py

# Conector de sql recomendado na documentação oficial.
from turtle import color
import pyodbc
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles.numbers import NumberFormatList
from openpyxl.styles import Font, Fill
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

def createWorkBook(nomeR, regional, os_por_regional_mes):
    if('BOLSAO/PARANAIBA' in nomeR):
        createSheetLocal('PNB', regional, os_por_regional_mes, nomeR)
    elif('GRANDE DOURADOS' in nomeR):
        createSheetLocal('DOS', regional, os_por_regional_mes, nomeR)
    elif('NORTE' in nomeR):
        createSheetLocal('CXM', regional, os_por_regional_mes, nomeR)
    elif('JIM' in nomeR):
        createSheetLocal('JIM', regional, os_por_regional_mes, nomeR)
    elif('CONE-SUL' in nomeR):
        createSheetLocal('NVR', regional, os_por_regional_mes, nomeR)
    elif('SUL/FRONTEIRA'in nomeR):
        createSheetLocal('PPR', regional, os_por_regional_mes, nomeR)
    elif('LESTE' in nomeR):
        createSheetLocal('NDI', regional, os_por_regional_mes, nomeR)
    elif('PANTANAL/CORUMBA' in nomeR):
        createSheetLocal('CMA', regional, os_por_regional_mes, nomeR)
    elif('PANTANAL/AQUIDAUANA' in nomeR):
        createSheetLocal('AUA', regional, os_por_regional_mes, nomeR)
    elif('BOLSAO/TRES LAGOAS' in nomeR):
        createSheetLocal('TLS', regional, os_por_regional_mes, nomeR)

def createSheetLocal(nomeR, regional, os_por_regional_mes, nomeRegional):
    wb = Workbook()

    col = [
        "DATA",
        "RDA",
        "RDA - Oneroso Terceiros",
        "Crescimento RDA",
        "OS Implantação de RDA",
        "RCE - Sanesul",
        "RCE - Oneroso Terceiros",
        "RCE - Oneroso Ms-Pantanal",
        "Crescimento RCE",
        "ECON",
        "PENDENTES - DDIF",
    ]
    com = [
        "DATA - Mês e Ano",
        "RDA - Rede de Distribuição de Água - Extensão de Rede de água com status da rede = ""existente""",
        "RDA - Oneroso Terceiros - Rede de Distribuição de Água Terceiros - Extensão de Rede de água com status da rede = ""Oneroso Terceiros""",
        "Crescimento RDA - Variação do Total de RDA em relação ao período anterior",
        "OS Implantação de RDA - Ordens de Serviço para Implantação de RDA",
        "RCE - Sanesul - Rede coletora de Esgoto Sanesul - Extensão de Rede de Esgoto com status da rede = ""existente""",
        "RCE - Oneroso Terceiros - Rede coletora de Esgoto Terceiros - Extensão de Rede de Esgoto com status da rede = ""Oneroso Terceiros""",
        "RCE - Oneroso MS-Pantanal - Rede coletora de Esgoto MS-Pantanal - Extensão de Rede de Esgoto com status da rede = ""Oneroso MS-Pantanal""",
        "Crescimento RCE - Variação do Total de RCE em relação ao período anterior",
        "ECON - Economia - Total de Matriculas Georreferenciadas",
        "PENDENTES - DDIF - Diferença entre o Total de Matriculas Cadastrais da Comercial e o Total de Matriculas Georreferenciados ou seja matriculas sem georreferenciamento, ou seja, pendentes de georreferenciamento.",
    ]
    groupLocal = regional.groupby('LOCALIDADE', sort=True)

    for nomeL, localidade in groupLocal:
        ws = wb.create_sheet(nomeL, -1)
        ws.append(col)

        for idx, texto_comentario in enumerate(com, start=1):
            comment = Comment(texto_comentario, "INFO", width=300, height=100)
            cell = ws.cell(row=1, column=idx)
            cell.comment = comment
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell_value = ws.cell(row=1, column=idx).value
            if cell_value:
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = len(str(cell_value)) + 5

        localidade = localidade.sort_values(by=['REFERENCIA', 'DATA'], kind='mergesort')

        rda_anterior = None
        rce_anterior = None

        for _, row in localidade.iterrows():
            referencia = row['REFERENCIA']
            if isinstance(referencia, str):
                referencia = datetime.strptime(referencia, "%m/%Y")
            else:
                referencia = pd.to_datetime(referencia).to_pydatetime()

            mes_key = referencia.strftime("%Y-%m")
            total_os = os_por_regional_mes.get((nomeRegional.strip(), mes_key), 0)

            rda_atual = (row.get('RDA', 0) or 0) + (row.get('RDA - Oneroso Terceiros', 0) or 0)
            rce_atual = (row.get('RCE - Sanesul', 0) or 0) + (row.get('RCE - Oneroso Terceiros', 0) or 0) + (row.get('RCE - Oneroso Ms-Pantanal', 0) or 0)

            if rda_anterior is None or rda_anterior == 0:
                crescimento_rda = 0
            else:
                crescimento_rda = (rda_atual - rda_anterior) / rda_anterior

            if rce_anterior is None or rce_anterior == 0:
                crescimento_rce = 0
            else:
                crescimento_rce = (rce_atual - rce_anterior) / rce_anterior

            nova_linha = [
                row['REFERENCIA'],
                row.get('RDA', 0) or 0,
                row.get('RDA - Oneroso Terceiros', 0) or 0,
                crescimento_rda,
                total_os,
                row.get('RCE - Sanesul', 0) or 0,
                row.get('RCE - Oneroso Terceiros', 0) or 0,
                row.get('RCE - Oneroso Ms-Pantanal', 0) or 0,
                crescimento_rce,
                row.get('ECON', 0) or 0,
                row.get('PENDENTES', 0) or 0,
            ]
            ws.append(nova_linha)

            rda_anterior = rda_atual
            rce_anterior = rce_atual

        for coluna in ["D", "I"]:
            for cell in ws[coluna][1:]:
                cell.number_format = "0.000%"

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            row[0].number_format = "mmm-yy"

        nLinhas = len(localidade)
        if nLinhas > 0:
            saldo = str(nLinhas + 2)
            ws["A" + saldo] = "Saldo = "
            ws["A" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

            ws["B" + saldo].value = ws["B" + str(nLinhas + 1)].value - ws["B2"].value
            ws["B" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

            ws["F" + saldo].value = ws["F" + str(nLinhas + 1)].value - ws["F2"].value
            ws["F" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
            ws["G" + saldo].value = ws["G" + str(nLinhas + 1)].value - ws["G2"].value
            ws["G" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

            ws["H" + saldo].value = ws["H" + str(nLinhas + 1)].value - ws["H2"].value
            ws["H" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

            ws["J" + saldo].value = ws["J" + str(nLinhas + 1)].value - ws["J2"].value
            ws["J" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

    wb.save('registro' + nomeR + '.xlsx')

def plotGraph(ws, max_row, nomeL):

    # Gráfico RDA e RCE
    chart1 = ScatterChart()
    # Gráfico Econ e DIFF
    chart2 = ScatterChart()

    xvalues = Reference(ws, min_col=1, min_row=2, max_row=max_row)

    yvalues = Reference(ws, min_col=2, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="RDA")
    chart1.series.append(series)

    yvalues = Reference(ws, min_col=6, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="RCE - Sanesul")
    chart1.series.append(series)

    yvalues = Reference(ws, min_col=7, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="RCE - Oneroso \n Terceiros")
    chart1.series.append(series)

    yvalues = Reference(ws, min_col=8, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="RCE - Oneroso \n Ms-Pantanal")
    chart1.series.append(series)

    yvalues = Reference(ws, min_col=10, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="ECON")
    chart2.series.append(series)

    yvalues = Reference(ws, min_col=11, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="PEND")
    chart2.series.append(series)

    chart1.title = nomeL + "--RDA e RCE"
    chart1.width = 16.5
    ws.add_chart(chart1, "M3")
    chart2.title = nomeL + "--ECON e PENDENTES"
    ws.add_chart(chart2, "M18")

if __name__ == '__main__':
    con = pyodbc.connect(
        'DRIVER={SQL Server};SERVER=10.100.100.48\\SCI;PORT=1433;DATABASE=SCI;Trusted_Connection=yes;')
    con_os = pyodbc.connect(
        'DRIVER={SQL Server};SERVER=10.100.100.48\\SCI;PORT=1433;DATABASE=SCI;Trusted_Connection=yes;')

    conHistorico = '''select REGIONAL, LOCAL as LOCALIDADE, REFERENCIA, COALESCE(RDA,0) AS RDA,
    COALESCE(RDA_TERCEIRO,0) AS [RDA - Oneroso Terceiros], COALESCE(RCE,0) AS [RCE - Sanesul],
    COALESCE(RCE_TERCEIRO,0) AS [RCE - Oneroso Terceiros], COALESCE(RCE_MSP,0) AS [RCE - Oneroso Ms-Pantanal],
    COALESCE(ECON,0) AS ECON, COALESCE(DDIFF,0) AS PENDENTES, DATAGERACAO as DATA
    from dbo.historico_resultado ORDER BY REGIONAL, LOCAL, DATAGERACAO asc'''
    conResultado = '''WITH CTE_UltimaDataPorMes AS (
        SELECT REGIONAL, LOCAL AS LOCALIDADE, REFERENCIA, RDA, COALESCE(RDA_TERCEIRO,0) AS [RDA - Oneroso Terceiros],
        COALESCE(RCE,0) AS [RCE - Sanesul], COALESCE(RCE_TERCEIRO,0) AS [RCE - Oneroso Terceiros],
        COALESCE(RCE_MSP,0) AS [RCE - Oneroso Ms-Pantanal], COALESCE(ECON,0) AS ECON, COALESCE(DDIFF,0) AS PENDENTES,
        DATAGERACAO AS DATA, ROW_NUMBER() OVER (PARTITION BY LOCAL, YEAR(DATAGERACAO), MONTH(DATAGERACAO) ORDER BY DATAGERACAO DESC) AS RN
        FROM dbo.resultado)
        SELECT REGIONAL, LOCALIDADE, REFERENCIA, RDA, [RDA - Oneroso Terceiros], [RCE - Sanesul], [RCE - Oneroso Terceiros], [RCE - Oneroso Ms-Pantanal], ECON, PENDENTES, DATA
        FROM CTE_UltimaDataPorMes WHERE RN = 1 ORDER BY REGIONAL, LOCALIDADE, DATA;'''

    dfHistorico = pd.read_sql(conHistorico, con)
    dfResultado = pd.read_sql(conResultado, con)

    dfgroup = dfResultado.groupby(["LOCALIDADE", "REFERENCIA"])
    dfResultado = dfgroup.tail(1)
    df = pd.concat([dfHistorico, dfResultado], ignore_index=True, sort=False)

    df['REGIONAL'] = df['REGIONAL'].str.strip().str.replace(r'\s+', ' ', regex=True)
    df = df.sort_values(by=['REGIONAL', 'LOCALIDADE'])

    os_por_regional_mes = {}
    sql_esquemas = "SELECT DISTINCT ESQUEMA, REGIONAL FROM historico_resultado"
    cursor = con.cursor()
    cursor_os = con_os.cursor()
    cursor.execute(sql_esquemas)

    regionais_esquemas = {}
    for esquema, nome_regional in cursor.fetchall():
        regionais_esquemas.setdefault(nome_regional.strip(), []).append(str(esquema))

    for regional_name, esquemas in regionais_esquemas.items():
        for esquema in esquemas:
            sql = f"""
            SELECT FORMAT(databaixa,'yyyy-MM') as mes, COUNT(*) as total
            FROM [{esquema}].[os_com_orfa]
            WHERE grupo = '57'
            GROUP BY FORMAT(databaixa,'yyyy-MM')
            """
            try:
                cursor_os.execute(sql)
                for mes, total in cursor_os.fetchall():
                    chave = (regional_name, mes)
                    os_por_regional_mes[chave] = os_por_regional_mes.get(chave, 0) + total
            except Exception as e:
                print(f"Erro no esquema {esquema}: {e}")

    groupRegional = df.groupby("REGIONAL")
    for nomeR, regional in groupRegional:
        createWorkBook(nomeR, regional, os_por_regional_mes)
