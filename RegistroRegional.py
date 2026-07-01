# pyinstaller RegistroRegional.py --onefile --noconsole

import os
import pyodbc
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles.numbers import NumberFormatList
from openpyxl.styles import Font, Fill
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ESTADO_DIR = os.path.join(BASE_DIR, 'estado')
os.makedirs(ESTADO_DIR, exist_ok=True)


def plotGraph(ws, max_row, nomeL):

    # Gráfico RDA e RCE
    chart1 = ScatterChart()
    # Gráfico Econ e DIFF
    chart2 = ScatterChart()

    xvalues = Reference(ws, min_col=1, min_row=2, max_row=max_row)

    yvalues = Reference(ws, min_col=2, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="RDA")
    chart1.series.append(series)

    yvalues = Reference(ws, min_col=6, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="RCE - Sanesul")
    chart1.series.append(series)

    yvalues = Reference(ws, min_col=7, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="RCE - Oneroso \n Terceiros")
    chart1.series.append(series)

    yvalues = Reference(ws, min_col=8, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="RCE - Oneroso \n Ms-Pantanal")
    chart1.series.append(series)

    yvalues = Reference(ws, min_col=10, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="ECON")
    chart2.series.append(series)

    yvalues = Reference(ws, min_col=11, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title="PEND")
    chart2.series.append(series)

    chart1.title = nomeL + "--RDA e RCE"
    chart1.width = 16.5
    ws.add_chart(chart1, "M3")
    chart2.title = nomeL + "--ECON e PENDENTES"
    ws.add_chart(chart2, "M18")


def nome(nomeR):
    if "BOLSAO/PARANAIBA" in nomeR:
        return "PNB"
    elif "GRANDE DOURADOS" in nomeR:
        return "DOS"
    elif "NORTE" in nomeR:
        return "CXM"
    elif "JIM" in nomeR:
        return "JIM"
    elif "CONE-SUL" in nomeR:
        return "NVR"
    elif "SUL/FRONTEIRA" in nomeR:
        return "PPR"
    elif "LESTE" in nomeR:
        return "NDI"
    elif "PANTANAL/CORUMBA" in nomeR:
        return "CMA"
    elif "PANTANAL/AQUIDAUANA" in nomeR:
        return "AUA"
    elif "BOLSAO/TRES LAGOAS" in nomeR:
        return "TLS"


if __name__ == "__main__":
    # Conectando com o banco de dados
    con = pyodbc.connect(
        "DRIVER={SQL Server};"
        "SERVER=10.100.100.48\\SCI;"
        "PORT=1433;"
        "DATABASE=SCI;"
        "Trusted_Connection=yes;"
    )
    con_os = pyodbc.connect(
    "DRIVER={SQL Server};"
    "SERVER=10.100.100.48\\SCI;"
    "DATABASE=SCI;"
    "Trusted_Connection=yes;"
    )

    con
    # Consultando a tabela do banco
    conRegional = "select distinct(regional) from resultado;"
    cursor = con.cursor()
    cursor_os = con_os.cursor()
    cursor.execute(conRegional)
    os_por_regional_mes = {}
    regionais = cursor.fetchall()
    wb = Workbook()
    conResultado = """select regional, max(datageracao) as datageracao, referencia as DATA,
            COALESCE(sum(RDA),0) as RDA, COALESCE(SUM(RDA_TERCEIRO),0) AS [RDA - Oneroso Terceiros],
            COALESCE(sum(RCE),0) as [RCE - Sanesul], COALESCE(sum(RCE_TERCEIRO),0) AS [RCE - Oneroso Terceiros],COALESCE(sum(RCE_MSP),0) AS [RCE - Oneroso MS-Pantanal],
            COALESCE(sum(ECON),0) as ECON, COALESCE(sum(DDIFF),0) as PEND	
            from historico_resultado
            where regional = ? group by regional, referencia, RCE_TERCEIRO, RCE_MSP
			union
            select regional, max(datageracao) as datageracao, referencia as DATA,
            COALESCE(sum(RDA),0) as RDA, COALESCE(SUM(RDA_TERCEIRO),0) AS [RDA - Oneroso Terceiros],
            COALESCE(sum(RCE),0) as [RCE - Sanesul], COALESCE(SUM(RCE_TERCEIRO),0) AS [RCE - Oneroso Terceiros], COALESCE(SUM(RCE_MSP),0) AS [RCE - Oneroso MS-Pantanal],
            COALESCE(sum(ECON),0) as ECON, COALESCE(sum(DDIFF),0) as PEND
            from resultado
            where datageracao in (select max(DATAGERACAO) as DATAGERACAO from resultado group by referencia, local)
			and regional = ?  group by regional, referencia order by regional, datageracao;"""
    
    sql_esquemas = """
    SELECT DISTINCT ESQUEMA, REGIONAL
    FROM historico_resultado
    """

    cursor.execute(sql_esquemas)

    regionais_esquemas = {}

    for esquema, nome_regional in cursor.fetchall():
        regionais_esquemas.setdefault(
            nome_regional.strip(),
            []
        ).append(str(esquema))

    for regional, esquemas in regionais_esquemas.items():
        for esquema in esquemas:

            sql = f"""
            SELECT
                FORMAT(databaixa,'yyyy-MM') as mes,
                COUNT(*) as total
            FROM [{esquema}].[os_com_orfa]
            WHERE grupo = '57'
            GROUP BY FORMAT(databaixa,'yyyy-MM')
            """

            try:
                cursor_os.execute(sql)

                for mes, total in cursor_os.fetchall():

                    chave = (regional, mes)

                    os_por_regional_mes[chave] = (
                        os_por_regional_mes.get(chave, 0)
                        + total
                    )

            except Exception as e:
                print(f"Erro no esquema {esquema}: {e}")

    for r in regionais:
        ws = wb.create_sheet(nome(r[0]), -1)
        col = [
            "DATA",
            "RDA",
            "RDA - Oneroso Terceiros",
            "Crescimento RDA",
            "OS Implantação de RDA",
            "RCE - Sanesul",
            "RCE - Oneroso Terceiros",
            "RCE - Oneroso Ms-Pantanal",
            "Crescimento RCE",
            "ECON",
            "PENDENTES - DDIF",
        ]
        com = [
            "DATA - Mês e Ano",
            "RDA - Rede de Distribuição de Água - Extensão de Rede de água com status da rede = ""existente""",
            "RDA - Oneroso Terceiros - Rede de Distribuição de Água Terceiros - Extensão de Rede de água com status da rede = ""Oneroso Terceiros""",
            "Crescimento RDA - Variação do Total de RDA em relação ao período anterior",
            "OS Implantação de RDA - Ordens de Serviço para Implantação de RDA",
            "RCE - Sanesul - Rede coletora de Esgoto Sanesul - Extensão de Rede de Esgoto com status da rede = ""existente""",
            "RCE - Oneroso Terceiros - Rede coletora de Esgoto Terceiros - Extensão de Rede de Esgoto com status da rede = ""Oneroso Terceiros""",
            "RCE - Oneroso MS-Pantanal - Rede coletora de Esgoto MS-Pantanal - Extensão de Rede de Esgoto com status da rede = ""Oneroso MS-Pantanal""",
            "Crescimento RCE - Variação do Total de RCE em relação ao período anterior",
            "ECON - Economia - Total de Matriculas Georreferenciadas",
            "PENDENTES - DDIF - Diferença entre o Total de Matriculas Cadastrais da Comercial e o Total de Matriculas Georreferenciados ou seja matriculas sem georreferenciamento, ou seja, pendentes de georreferenciamento.",
        ]
        ws.append(col)
        # Inclui comentários descritivos nas células de cabeçalho
        for idx, texto_comentario in enumerate(com, start=1):
            comment = Comment(texto_comentario, "INFO", width=300, height=100)
            cell = ws.cell(row=1, column=idx)
            cell.comment = comment
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Ajusta a largura da coluna com base no comprimento do texto do cabeçalho
            cell_value = ws.cell(row=1, column=idx).value
            if cell_value:
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = len(str(cell_value)) + 5
        nLinhas = 0
        cursor.execute(conResultado, r[0], r[0])
        row = cursor.fetchone()
        print("Regional:", r[0])
        print("Primeira linha:", row)
        print("Regional:", r[0], "Linhas:", nLinhas)
        anterior = 0
        rda_anterior = None
        rce_anterior = None

        while row != None:
            referencia = datetime.strptime(row[2], "%m/%Y")

            if referencia != anterior:

                chave = (
                    r[0].strip(),
                    referencia.strftime("%Y-%m")
                )

                total_os = os_por_regional_mes.get(chave, 0)
                
                nLinhas += 1

                rda_atual = (
                    (row[3] or 0) +
                    (row[4] or 0)
                )
                rce_atual = (
                    (row[5] or 0) +
                    (row[6] or 0) +
                    (row[7] or 0)
                )

                if rda_anterior is None or rda_anterior == 0:
                    crescimento_rda = 0
                else:
                    crescimento_rda = (
                        (rda_atual - rda_anterior) / rda_anterior
                    )

                if rce_anterior is None or rce_anterior == 0:
                    crescimento_rce = 0
                else:
                    crescimento_rce = (
                        (rce_atual - rce_anterior) / rce_anterior
                    )


                nova_linha = [
                    row[2],                 # DATA
                    row[3],                 # RDA
                    row[4],                 # RDA - Oneroso Terceiros
                    crescimento_rda,        # Crescimento RDA
                    total_os,               # OS Implantação de RDA
                    row[5],                 # RCE - Sanesul
                    row[6],                 # RCE - Oneroso Terceiros
                    row[7],                 # RCE - Oneroso MS-Pantanal
                    crescimento_rce,        # Crescimento RCE
                    row[8],                 # ECON
                    row[9],                 # PENDENTES
                ]

                ws.append(nova_linha)

                rda_anterior = rda_atual
                rce_anterior = rce_atual

            anterior = referencia
            row = cursor.fetchone()

        for coluna in ["D", "I"]:
            for cell in ws[coluna]:
                cell.number_format = "0.000%"

        for cell in ws["A"]:
            cell.number_format = "mmm-yy"
        plotGraph(ws, nLinhas + 1, r[0])
        saldo = str(nLinhas + 2)
        ws["A" + saldo] = "Saldo = "
        ws["A" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

        print("nLinhas =", nLinhas)
        print("B2 =", ws["B2"].value)
        print("B" + str(nLinhas + 1), "=", ws["B" + str(nLinhas + 1)].value)

        ws["B" + saldo].value = ws["B" + str(nLinhas + 1)].value - ws["B2"].value
        ws["B" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

        ws["F" + saldo].value = ws["F" + str(nLinhas + 1)].value - ws["F2"].value
        ws["F" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")
        ws["G" + saldo].value = ws["G" + str(nLinhas + 1)].value - ws["G2"].value
        ws["G" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

        ws["H" + saldo].value = ws["H" + str(nLinhas + 1)].value - ws["H2"].value
        ws["H" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

        ws["J" + saldo].value = ws["J" + str(nLinhas + 1)].value - ws["J2"].value
        ws["J" + saldo].fill = PatternFill("solid", fgColor="00FFFF00")

    wb.save(os.path.join(ESTADO_DIR, "registroRegional.xlsx"))
